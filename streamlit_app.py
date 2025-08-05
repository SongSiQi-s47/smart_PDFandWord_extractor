# streamlit_app.py
import streamlit as st
import os
import tempfile
import pandas as pd
from pdf_extractorV2_2 import PDFWordTableExtractor

def main():
    st.set_page_config(
        page_title="智能表格提取工具",
        page_icon="📋",
        layout="wide",
        initial_sidebar_state="collapsed"
    )
    
    # 自定义CSS样式
    st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    .upload-area {
        border: 3px dashed #1f77b4;
        border-radius: 15px;
        padding: 3rem;
        text-align: center;
        background: linear-gradient(135deg, #f0f8ff 0%, #e6f3ff 100%);
        margin: 2rem 0;
    }
    .success-box {
        background: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .error-box {
        background: #f8d7da;
        border: 1px solid #f5c6cb;
        border-radius: 10px;
        padding: 1rem;
        margin: 1rem 0;
    }
    </style>
    """, unsafe_allow_html=True)
    
    st.markdown('<h1 class="main-header">📋 智能表格提取工具</h1>', unsafe_allow_html=True)
    
    # 功能介绍
    with st.expander("ℹ️ 功能介绍", expanded=False):
        st.markdown("""
        **功能特点：**
        - 📄 智能识别PDF和Word文档中的表格
        - 📊 自动提取分项报价表数据
        - 🏗️ 支持多级模块结构识别
        - 📥 一键导出Excel文件
        - 🚀 零安装，即开即用
        
        **支持格式：**
        - PDF文件 (.pdf)
        - Word文档 (.docx)
        
        **使用说明：**
        1. 输入编号样例（用于识别模块层级）
        2. 设置表头映射（Word合同文件）
        3. 上传包含"分项报价表"的PDF或Word文件
        4. 系统自动识别表格结构
        5. 点击"开始提取"进行处理
        6. 下载Excel格式的结果文件
        """)
    
    st.markdown("---")
    
    # 编号样例输入区域
    st.subheader("🔢 编号样例设置")
    st.markdown("**请根据你的文档格式输入编号样例：**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        lvl1_sample = st.text_input(
            "一级模块编号样例",
            placeholder="如：9.1.3.4 或 （一）",
            help="输入文档中一级模块的编号格式"
        )
        
        lvl2_sample = st.text_input(
            "二级模块编号样例",
            placeholder="如：9.1.3.4.1 或 （二）",
            help="输入文档中二级模块的编号格式（可选）"
        )
    
    with col2:
        lvl3_sample = st.text_input(
            "三级模块编号样例",
            placeholder="如：1",
            help="输入文档中三级模块的编号格式（可选）"
        )
        
        end_sample = st.text_input(
            "终止编号样例",
            placeholder="如：结束 或 附录",
            help="输入遇到该编号时停止提取（可选）"
        )
    
    st.markdown("---")
    
    # 表头设置区域（仅对Word合同文件显示）
    st.subheader("📋 表头设置（Word合同文件）")
    st.markdown("**如果处理Word合同文件，请设置表头映射：**")
    
    col1, col2 = st.columns(2)
    
    with col1:
        lvl1_header = st.text_input(
            "一级模块名称对应表头",
            placeholder="如：功能模块、模块名称等",
            help="Word文档中对应一级模块名称的列标题"
        )
        
        lvl2_header = st.text_input(
            "二级模块名称对应表头",
            placeholder="如：功能子项、子模块等",
            help="Word文档中对应二级模块名称的列标题"
        )
    
    with col2:
        lvl3_header = st.text_input(
            "三级模块名称对应表头",
            placeholder="如：三级模块、子项等",
            help="Word文档中对应三级模块名称的列标题"
        )
        
        desc_header = st.text_input(
            "合同描述对应表头",
            placeholder="如：功能描述、描述、备注等",
            help="Word文档中对应合同描述的列标题"
        )
    
    # 创建表头映射
    custom_headers = {}
    if lvl1_header:
        custom_headers[lvl1_header] = '一级模块名称'
    if lvl2_header:
        custom_headers[lvl2_header] = '二级模块名称'
    if lvl3_header:
        custom_headers[lvl3_header] = '三级模块名称'
    if desc_header:
        custom_headers[desc_header] = '合同描述'
    
    # 如果没有设置自定义表头，为Word合同文件提供默认映射
    if not custom_headers:
        custom_headers = {
            '功能模块': '一级模块名称',
            '功能子项': '二级模块名称', 
            '三级模块': '三级模块名称',
            '功能描述': '合同描述'
        }
    
    st.markdown("---")
    
    # 文件上传区域
    st.markdown('<div class="upload-area">', unsafe_allow_html=True)
    st.markdown("### 📁 文件上传")
    st.markdown("**支持拖拽上传多个文件**")
    
    uploaded_files = st.file_uploader(
        "选择PDF或Word文件",
        type=['pdf', 'docx'],
        accept_multiple_files=True,
        help="可以同时选择多个文件，支持拖拽上传"
    )
    st.markdown('</div>', unsafe_allow_html=True)
    
    if uploaded_files:
        # 显示上传的文件
        st.markdown("### 📋 已上传文件")
        for i, file in enumerate(uploaded_files):
            file_size = len(file.getvalue()) / 1024  # KB
            st.write(f"**{i+1}.** {file.name} ({file_size:.1f} KB)")
        
        # 处理按钮
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🚀 开始提取", type="primary", use_container_width=True):
                if not lvl1_sample:
                    st.error("❌ 请至少输入一级模块编号样例！")
                else:
                    process_files(uploaded_files, lvl1_sample, lvl2_sample, lvl3_sample, end_sample, custom_headers)
    
    else:
        st.info("👆 请上传PDF或Word文件开始提取")
        
        # 添加示例说明
        st.markdown("### 💡 使用提示")
        st.markdown("""
        - 确保文件包含"分项报价表"相关内容
        - 支持标书和合同文件的自动识别
        - 系统会自动识别表格结构和模块层级
        - 提取结果将保存为Excel格式
        """)

def process_files(uploaded_files, lvl1_sample, lvl2_sample, lvl3_sample, end_sample, custom_headers=None):
    """处理上传的文件"""
    extractor = PDFWordTableExtractor()
    
    # 设置自定义表头
    if custom_headers:
        extractor.custom_headers = custom_headers
    
    all_data = []
    
    # 创建进度条
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    for i, uploaded_file in enumerate(uploaded_files):
        status_text.text(f"正在处理: {uploaded_file.name}")
        
        # 修改这里：使用正确的文件扩展名
        file_extension = os.path.splitext(uploaded_file.name)[1]  # 获取 .pdf 或 .docx
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_extension) as tmp_file:
            tmp_file.write(uploaded_file.getvalue())
            tmp_file_path = tmp_file.name
        
        try:
            # 确定文件类型
            file_type = "pdf" if uploaded_file.type == "application/pdf" else "docx"
            
            # 修改：传递编号样例给提取函数
            if file_type == "pdf":
                # PDF文件使用编号样例提取
                if lvl1_sample:  # 确保有编号样例
                    data = extractor.extract_tables_from_pdf_bid_with_samples(tmp_file_path, lvl1_sample, lvl2_sample, lvl3_sample, end_sample)
                else:
                    # 如果没有编号样例，尝试自动识别
                    data = extractor.extract_tables(tmp_file_path, file_type)
            else:
                # Word文件使用原有的提取方法
                data = extract_tables_with_samples(extractor, tmp_file_path, file_type, 
                                                lvl1_sample, lvl2_sample, lvl3_sample, end_sample)
            
            if data:
                all_data.extend(data)
                st.markdown(f'<div class="success-box">✅ {uploaded_file.name}: 提取到 {len(data)} 条记录</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div class="error-box">⚠️ {uploaded_file.name}: 未提取到数据</div>', unsafe_allow_html=True)
        
        except Exception as e:
            st.markdown(f'<div class="error-box">❌ {uploaded_file.name}: 处理失败 - {str(e)}</div>', unsafe_allow_html=True)
        
        finally:
            # 清理临时文件
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)
        
        # 更新进度
        progress_bar.progress((i + 1) / len(uploaded_files))
    
    # 显示结果
    if all_data:
        st.success(f"🎉 提取完成！共 {len(all_data)} 条记录")
        
        # 创建DataFrame并确保列顺序正确
        df = pd.DataFrame(all_data)
        
        # 确保所有必需的列都存在
        required_columns = ['一级模块名称', '二级模块名称', '三级模块名称', '标书描述', '合同描述', '来源文件']
        for col in required_columns:
            if col not in df.columns:
                df[col] = ''  # 添加缺失的列，填充空字符串
        
        # 按正确顺序排列列
        column_order = ['一级模块名称', '二级模块名称', '三级模块名称', '标书描述', '合同描述', '来源文件']
        df = df[column_order]
        
        # 显示数据预览
        st.subheader("📊 数据预览")
        st.dataframe(df.head(10), use_container_width=True)
        
        # 下载按钮
        st.subheader("📥 下载结果")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # CSV下载
            csv_data = df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📄 下载CSV文件",
                data=csv_data,
                file_name="提取结果.csv",
                mime="text/csv",
                use_container_width=True
            )
        
        with col2:
            # Excel下载 - 添加格式化
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # 创建Excel文件
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='提取结果', index=False)
                worksheet = writer.sheets['提取结果']
                
                # 设置列宽
                column_order = ['一级模块名称', '二级模块名称', '三级模块名称', '标书描述', '合同描述', '来源文件']
                column_widths = {
                    'A': 15, 'B': 15, 'C': 15, 'D': 45, 'E': 45, 'F': 10
                }
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 设置表头样式
                header_font = Font(bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center')
                for col in range(1, len(column_order) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                # 设置数据行样式和行高
                data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                for row in range(2, len(df) + 2):
                    # 计算每行最大字符数
                    max_chars = 0
                    for col in range(1, len(column_order) + 1):
                        cell_value = str(worksheet.cell(row=row, column=col).value or '')
                        lines = cell_value.split('\n')
                        for line in lines:
                            line_chars = len(line)
                            if line_chars > 30:
                                needed_lines = (line_chars // 30) + 1
                                max_chars = max(max_chars, needed_lines * 30)
                            else:
                                max_chars = max(max_chars, line_chars)
                    
                    # 计算行高
                    estimated_lines = max(1, (max_chars // 30) + 1)
                    row_height = max(20, estimated_lines * 18 + 10)
                    worksheet.row_dimensions[row].height = row_height
                    
                    # 设置单元格对齐方式
                    for col in range(1, len(column_order) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = data_alignment
                
                # 设置边框
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(column_order)):
                    for cell in row:
                        cell.border = thin_border
            
            excel_data = excel_buffer.getvalue()
            
            st.download_button(
                label="📊 下载Excel文件",
                data=excel_data,
                file_name="提取结果.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.error("❌ 未提取到任何数据")

def extract_tables_with_samples(extractor, file_path, file_type, lvl1_sample, lvl2_sample, lvl3_sample, end_sample):
    """使用编号样例提取表格"""
    if file_type == "pdf":
        return extractor.extract_tables_with_samples(file_path, file_type, lvl1_sample, lvl2_sample, lvl3_sample, end_sample)
    elif file_type == "docx":
        # Word文件使用原有的提取方法
        if "合同" in file_path:
            return extractor.extract_tables_from_word_contract(file_path)
        else:
            return extractor.extract_tables_from_word_bid(file_path)
    else:
        return []

if __name__ == "__main__":
    main()